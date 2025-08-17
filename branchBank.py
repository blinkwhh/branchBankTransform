import os
from openpyxl import Workbook
from openpyxl.styles import Font

def ip_to_int(ip_str):
    """将点分十进制IP地址转换为32位整数"""
    parts = ip_str.split('.')
    return (int(parts[0]) << 24) | (int(parts[1]) << 16) | \
           (int(parts[2]) << 8) | int(parts[3])

def int_to_ip(ip_int):
    """将32位整数转换为点分十进制IP地址"""
    return f"{(ip_int >> 24) & 0xFF}.{(ip_int >> 16) & 0xFF}." \
           f"{(ip_int >> 8) & 0xFF}.{ip_int & 0xFF}"

def acl_contains_cidr(acl_str, cidr_str):
    """
    判断ACL（反掩码表示）是否包含指定的CIDR网段
    
    参数:
        acl_str: ACL字符串，格式如 "10.0.0.0 0.0.0.7"
        cidr_str: CIDR网段字符串，格式如 "10.0.0.0/30"
    
    返回:
        bool: 如果ACL包含CIDR网段则返回True，否则返回False
    """
    # 解析ACL
    acl_parts = acl_str.split()
    if len(acl_parts) != 2:
        raise ValueError("ACL格式错误，应为'<IP地址> <反掩码>'")
    
    acl_ip_str, wildcard_str = acl_parts
    acl_ip = ip_to_int(acl_ip_str)
    wildcard = ip_to_int(wildcard_str)
    
    # 计算ACL的起始和结束地址
    acl_start = acl_ip & ~wildcard  # 反掩码0位必须匹配
    acl_end = acl_ip | wildcard     # 反掩码1位可以是任意值
    
    # 解析CIDR
    cidr_parts = cidr_str.split('/')
    if len(cidr_parts) != 2:
        raise ValueError("CIDR格式错误，应为'<IP地址>/<前缀长度>'")
    
    cidr_ip_str, prefix_len = cidr_parts
    prefix_len = int(prefix_len)
    cidr_ip = ip_to_int(cidr_ip_str)
    
    # 计算CIDR的起始和结束地址
    mask = (0xFFFFFFFF << (32 - prefix_len)) & 0xFFFFFFFF
    cidr_start = cidr_ip & mask
    cidr_end = cidr_start | (0xFFFFFFFF ^ mask)
    
    # 判断CIDR是否完全在ACL范围内
    return acl_start <= cidr_start and acl_end >= cidr_end

def load_acl_from_file(filename):
    """从文件加载ACL列表"""
    acl_list = []
    try:
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                # 跳过空行和注释行
                if line and not line.startswith('#'):
                    acl_list.append(line)
    except FileNotFoundError:
        print(f"错误: 文件 {filename} 未找到")
    return acl_list

def load_cidr_from_file(filename):
    """从文件加载CIDR列表"""
    cidr_list = []
    try:
        with open(filename, 'r') as file:
            for line in file:
                line = line.strip()
                # 跳过空行和注释行
                if line and not line.startswith('#'):
                    cidr_list.append(line)
    except FileNotFoundError:
        print(f"错误: 文件 {filename} 未找到")
    return cidr_list

def export_to_excel(data, filename, title="未覆盖的CIDR列表"):
    """将数据导出到Excel文件"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "未覆盖CIDR"
    
    # 添加标题
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    
    # 添加表头
    ws['A3'] = "序号"
    ws['B3'] = "CIDR网段"
    ws['A3'].font = Font(bold=True)
    ws['B3'].font = Font(bold=True)
    
    # 添加数据
    for idx, cidr in enumerate(data, start=1):
        ws.append([idx, cidr])
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    
    # 保存文件
    wb.save(filename)
    print(f"Excel文件已生成: {os.path.abspath(filename)}")

def batch_verify_acl_coverage(acl_file, cidr_file, excel_output="uncovered_cidrs.xlsx"):
    """批量验证CIDR网段是否被ACL覆盖"""
    # 加载ACL和CIDR列表
    acl_list = load_acl_from_file(acl_file)
    cidr_list = load_cidr_from_file(cidr_file)
    
    if not acl_list:
        print("错误: 未加载任何ACL")
        return
    
    if not cidr_list:
        print("错误: 未加载任何CIDR")
        return
    
    print(f"已加载 {len(acl_list)} 条ACL规则")
    print(f"已加载 {len(cidr_list)} 条CIDR网段")
    print("=" * 60)
    
    # 统计结果
    covered_count = 0
    not_covered = []
    
    # 检查每个CIDR是否被至少一个ACL覆盖
    for cidr in cidr_list:
        covered = False
        for acl in acl_list:
            try:
                if acl_contains_cidr(acl, cidr):
                    covered = True
                    break  # 找到一个覆盖的ACL即可
            except ValueError as e:
                print(f"处理错误: CIDR '{cidr}' 与 ACL '{acl}': {e}")
                continue
        
        if covered:
            covered_count += 1
        else:
            not_covered.append(cidr)
            print(f"[未覆盖] {cidr}")
    
    # 输出统计结果
    print("=" * 60)
    print(f"覆盖统计:")
    print(f"总CIDR数量: {len(cidr_list)}")
    print(f"已覆盖数量: {covered_count}")
    print(f"未覆盖数量: {len(not_covered)}")
    
    # 输出未覆盖的CIDR列表
    if not_covered:
        export_to_excel(not_covered, excel_output)
        print(f"\n已导出 {len(not_covered)} 条未覆盖CIDR到Excel文件")
    else:
        print("\n所有CIDR网段都被ACL覆盖！")

# 主程序
if __name__ == "__main__":
    # 文件路径
    acl_file = "PRDOA_ACL.txt"
    cidr_file = "WN_DS_routes.txt"
    excel_output = "uncovered_cidrs.xlsx"
    
    # 执行批量验证
    batch_verify_acl_coverage(acl_file, cidr_file, excel_output)